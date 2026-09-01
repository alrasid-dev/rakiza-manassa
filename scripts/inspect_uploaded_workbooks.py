from pathlib import Path
import json
import sys
from openpyxl import load_workbook

files = [Path(p) for p in sys.argv[1:]]
result = []
for path in files:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in workbook.worksheets:
        rows = ws.iter_rows(values_only=True)
        sample = []
        for row in rows:
            values = [str(value).strip() if value is not None else None for value in row]
            if any(value not in (None, "") for value in values):
                sample.append(values[:20])
            if len(sample) >= 8:
                break
        sheets.append({"title": ws.title, "max_row": ws.max_row, "max_column": ws.max_column, "sample": sample})
    result.append({"file": path.name, "sheets": sheets})
print(json.dumps(result, ensure_ascii=False, indent=2))
